# =============================================================================
# PROGRAMA: Sistema de Asistencia - Grupo Bravvo S.A.C
# DESCRIPCION: Servidor Flask que implementa el Algoritmo de Cristian para
#              sincronizar la hora entre cliente y servidor al momento de
#              registrar la asistencia de empleados.
# ALGORITMO: Cristian (1989) - Sincronizacion de relojes distribuidos
# =============================================================================

from flask import Flask, render_template, request, jsonify, send_file
import time
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import cm

# Inicializacion de la aplicacion Flask
app = Flask(__name__)

# =============================================================================
# ALMACENAMIENTO EN MEMORIA
# Lista global que guarda todos los registros de asistencia de la sesion.
# En produccion se reemplazaria por una base de datos.
# =============================================================================
registros = []


# =============================================================================
# FUNCION: algoritmo_cristian
# DESCRIPCION: Implementa el calculo central del Algoritmo de Cristian.
#   - T0: momento en que el cliente envio la solicitud (timestamp del cliente)
#   - T_servidor: hora oficial leida en el servidor al recibir la solicitud
#   - T1: momento en que el cliente recibe la respuesta (calculado con RTT)
#   El cliente estima su hora correcta como: T_servidor + RTT/2
# PARAMETROS:
#   t0_cliente_ms  -- timestamp del cliente al enviar (en milisegundos)
#   t1_cliente_ms  -- timestamp del cliente al recibir (en milisegundos)
#   t_servidor_ms  -- hora del servidor al procesar (en milisegundos)
# RETORNA: dict con RTT, hora estimada y error de sincronizacion
# =============================================================================
def algoritmo_cristian(t0_cliente_ms, t1_cliente_ms, t_servidor_ms):

    # RTT = tiempo total de ida y vuelta de la solicitud
    rtt = t1_cliente_ms - t0_cliente_ms

    # Hora estimada del cliente segun Cristian: T_servidor + RTT/2
    # Asume que la latencia de ida es aproximadamente igual a la de vuelta
    t_estimado = t_servidor_ms + (rtt / 2)

    # Incertidumbre: no podemos saber exactamente cuanto tardo cada direccion
    incertidumbre = rtt / 2

    # Error entre la hora del cliente y la hora sincronizada
    error_ms = abs(t_estimado - t0_cliente_ms)

    return {
        "rtt_ms":           round(rtt, 2),
        "t_estimado_ms":    round(t_estimado, 2),
        "incertidumbre_ms": round(incertidumbre, 2),
        "error_ms":         round(error_ms, 2),
        "valido":           rtt < 5000  # rechaza si RTT > 5 segundos
    }


# =============================================================================
# RUTA: GET /
# DESCRIPCION: Sirve la pagina principal (formulario de asistencia)
# =============================================================================
@app.route("/")
def index():
    return render_template("index.html")


# =============================================================================
# RUTA: GET /admin
# DESCRIPCION: Panel del profesor/administrador con todos los registros
# =============================================================================
@app.route("/admin")
def admin():
    return render_template("admin.html")


# =============================================================================
# RUTA: GET /api/tiempo
# DESCRIPCION: Endpoint que el cliente consulta para obtener la hora del
#              servidor. Es el paso 1 del Algoritmo de Cristian: el cliente
#              envia T0 en la solicitud y recibe T_servidor en la respuesta.
# =============================================================================
@app.route("/api/tiempo")
def api_tiempo():

    # Leer la hora del servidor en este instante exacto
    t_servidor_ms = time.time() * 1000

    # T0 enviado por el cliente (para calcular RTT en el servidor si se desea)
    t0_str = request.args.get("t0", "0")

    return jsonify({
        "t_servidor_ms": round(t_servidor_ms, 2),
        "hora_legible":  datetime.fromtimestamp(t_servidor_ms / 1000).strftime("%I:%M:%S %p"),
        "fecha":         datetime.fromtimestamp(t_servidor_ms / 1000).strftime("%d/%m/%Y"),
        "t0_recibido":   float(t0_str)
    })


# =============================================================================
# RUTA: POST /api/registrar
# DESCRIPCION: Recibe el registro de asistencia del empleado junto con los
#              timestamps necesarios para ejecutar el Algoritmo de Cristian
#              y almacena el resultado.
# BODY JSON:
#   nombre         -- nombre del empleado
#   t0_ms          -- hora del cliente al enviar (ms)
#   t1_ms          -- hora del cliente al recibir respuesta (ms)
#   t_servidor_ms  -- hora del servidor recibida en /api/tiempo
#   hora_cliente   -- hora local del cliente (texto legible)
# =============================================================================
@app.route("/api/registrar", methods=["POST"])
def api_registrar():

    data = request.get_json()

    # Extraer datos del cuerpo de la solicitud
    nombre        = data.get("nombre", "").strip()
    t0_ms         = float(data.get("t0_ms", 0))
    t1_ms         = float(data.get("t1_ms", 0))
    t_servidor_ms = float(data.get("t_servidor_ms", 0))
    hora_cliente  = data.get("hora_cliente", "")

    # Validar que se envio un nombre
    if not nombre:
        return jsonify({"error": "Nombre requerido"}), 400

    # Ejecutar el Algoritmo de Cristian
    resultado = algoritmo_cristian(t0_ms, t1_ms, t_servidor_ms)

    # Construir la hora oficial sincronizada (legible)
    hora_oficial = datetime.fromtimestamp(
        resultado["t_estimado_ms"] / 1000
    ).strftime("%I:%M:%S %p")

    # Construir el registro completo
    registro = {
        "n":              len(registros) + 1,
        "nombre":         nombre,
        "hora_cliente":   hora_cliente,         # hora local (desfasada)
        "hora_oficial":   hora_oficial,          # hora sincronizada por Cristian
        "rtt_ms":         resultado["rtt_ms"],
        "error_ms":       resultado["error_ms"],
        "incertidumbre":  resultado["incertidumbre_ms"],
        "valido":         resultado["valido"],
        "estado":         "Validado" if resultado["valido"] else "RTT muy alto",
        "fecha":          datetime.now().strftime("%d/%m/%Y"),
        "timestamp":      time.time()
    }

    # Guardar en memoria
    registros.append(registro)

    return jsonify({
        "ok":           True,
        "rtt_ms":       resultado["rtt_ms"],
        "hora_oficial": hora_oficial,
        "hora_cliente": hora_cliente,
        "error_ms":     resultado["error_ms"],
        "estado":       registro["estado"],
        "valido":       resultado["valido"]
    })


# =============================================================================
# RUTA: GET /api/registros
# DESCRIPCION: Retorna todos los registros de asistencia (para el panel
#              del profesor). Permite filtrar por fecha.
# =============================================================================
@app.route("/api/registros")
def api_registros():
    return jsonify(registros)


# =============================================================================
# RUTA: POST /api/exportar/excel
# DESCRIPCION: Genera y descarga un archivo Excel con todos los registros
#              de asistencia del dia, con formato profesional.
# =============================================================================
@app.route("/api/exportar/excel")
def exportar_excel():

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencia Bravvo"

    # --- Encabezado principal ---
    ws.merge_cells("A1:H1")
    ws["A1"] = "GRUPO BRAVVO S.A.C — Registro de Asistencia"
    ws["A1"].font      = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill      = PatternFill("solid", fgColor="0D3B66")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:H2")
    ws["A2"] = f"Fecha: {datetime.now().strftime('%d/%m/%Y')}   |   Algoritmo de Cristian — Sincronización de Relojes"
    ws["A2"].font      = Font(italic=True, size=10)
    ws["A2"].alignment = Alignment(horizontal="center")

    # --- Cabeceras de columna ---
    headers = ["#", "Empleado", "Hora Cliente\n(desfasada)", "Hora Oficial\n(Cristian)",
               "RTT (ms)", "Error (ms)", "Incertidumbre", "Estado"]
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font      = Font(bold=True, color="FFFFFF")
        cell.fill      = PatternFill("solid", fgColor="00695C")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border    = border
    ws.row_dimensions[4].height = 32

    # --- Filas de datos ---
    for r, reg in enumerate(registros, 5):
        fill_color = "D4EDDA" if reg["valido"] else "F8D7DA"
        valores = [
            reg["n"], reg["nombre"], reg["hora_cliente"], reg["hora_oficial"],
            reg["rtt_ms"], reg["error_ms"], f"±{reg['incertidumbre']} ms", reg["estado"]
        ]
        for c, val in enumerate(valores, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.fill      = PatternFill("solid", fgColor=fill_color)
            cell.border    = border
            cell.alignment = Alignment(horizontal="center")

    # Ajuste de anchos (usar get_column_letter para evitar MergedCell)
    from openpyxl.utils import get_column_letter
    anchos = [5, 22, 16, 16, 10, 10, 16, 14]
    for i, ancho in enumerate(anchos, 1):
        ws.column_dimensions[get_column_letter(i)].width = ancho

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True,
                     download_name="asistencia_bravvo.xlsx")


# =============================================================================
# RUTA: GET /api/exportar/pdf
# DESCRIPCION: Genera y descarga un PDF con el reporte de asistencia.
# =============================================================================
@app.route("/api/exportar/pdf")
def exportar_pdf():

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            topMargin=1.5*cm, bottomMargin=1.5*cm,
                            leftMargin=1.5*cm, rightMargin=1.5*cm)
    story = []

    title_s = ParagraphStyle("t", fontSize=15, fontName="Helvetica-Bold",
                              textColor=colors.HexColor("#0D3B66"), spaceAfter=4)
    sub_s   = ParagraphStyle("s", fontSize=9,  fontName="Helvetica",
                              textColor=colors.HexColor("#546E7A"), spaceAfter=14)

    story.append(Paragraph("GRUPO BRAVVO S.A.C — Registro de Asistencia", title_s))
    story.append(Paragraph(
        f"Fecha: {datetime.now().strftime('%d/%m/%Y %I:%M %p')}   |   "
        f"Algoritmo de Cristian — Sincronización de Relojes   |   "
        f"Total registros: {len(registros)}", sub_s))

    # Tabla
    tabla_data = [["#", "Empleado", "Hora Cliente", "Hora Oficial (Cristian)",
                   "RTT (ms)", "Error (ms)", "Estado"]]
    row_colors_list = []

    for reg in registros:
        tabla_data.append([
            str(reg["n"]), reg["nombre"], reg["hora_cliente"],
            reg["hora_oficial"], str(reg["rtt_ms"]),
            str(reg["error_ms"]), reg["estado"]
        ])
        row_colors_list.append(
            colors.HexColor("#D4EDDA") if reg["valido"] else colors.HexColor("#F8D7DA")
        )

    style = TableStyle([
        ("BACKGROUND",  (0,0), (-1,0), colors.HexColor("#0D3B66")),
        ("TEXTCOLOR",   (0,0), (-1,0), colors.white),
        ("FONTNAME",    (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE",    (0,0), (-1,-1), 8),
        ("ALIGN",       (0,0), (-1,-1), "CENTER"),
        ("GRID",        (0,0), (-1,-1), 0.5, colors.HexColor("#CCCCCC")),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), row_colors_list),
    ])

    t = Table(tabla_data,
              colWidths=[1*cm, 4*cm, 2.8*cm, 3.5*cm, 2*cm, 2*cm, 2.2*cm])
    t.setStyle(style)
    story.append(t)

    doc.build(story)
    buf.seek(0)
    return send_file(buf, mimetype="application/pdf",
                     as_attachment=True,
                     download_name="asistencia_bravvo.pdf")


# =============================================================================
# RUTA: DELETE /api/limpiar
# DESCRIPCION: Limpia todos los registros de la sesion (uso del profesor)
# =============================================================================
@app.route("/api/limpiar", methods=["DELETE"])
def limpiar():
    registros.clear()
    return jsonify({"ok": True})


# =============================================================================
# PUNTO DE ENTRADA
# Para desarrollo local: python app.py
# Para Render: gunicorn usa la variable 'app' directamente
# =============================================================================


# =============================================================================
# RUTA: GET /ping
# DESCRIPCION: Endpoint liviano para mantener el servidor de Render despierto.
#              El cliente hace fetch a esta ruta cada 10 minutos para evitar
#              que Render duerma el servicio en el plan gratuito.
# =============================================================================
@app.route("/ping")
def ping():
    return jsonify({"ok": True, "ts": time.time()})
if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=5000)
